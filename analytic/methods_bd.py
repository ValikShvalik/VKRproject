from PyQt5.QtCore import QThread, pyqtSignal
import os, shutil, sqlite3, openpyxl, requests, subprocess, time
from Global import manual_widths, type_names
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


DB_PATH = "database/converter.db"
EXPORT_DIR = "database/export"
FIELD_MAPPING = {
    "Порядковый номер": "serial_number",
    "Время": "time",
    "Номер задачи": "task_number",
    "Тип диагностического сообщения": "message_type",
    "Длина бинарных данных": "data_length",
    "Бинарные данные": "data_blob",
    "Текстовое сообщение разработчику": "developer_note"
}

def get_all_files_from_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, file_name, added_at FROM xlsx_files")
    rows = cursor.fetchall()
    conn.close()
    return [{"id": row[0], "file_name": row[1], "added_at": row[2]} for row in rows]

def get_db_size_mb():
    if os.path.exists(DB_PATH):
        size_bytes = os.path.getsize(DB_PATH)
        return size_bytes / (1024 * 1024)
    return 0

def delete_selected_files(file_ids: list[int]):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    for file_id in file_ids:
        cursor.execute("DELETE FROM xlsx_files WHERE id = ?", (file_id,))
    conn.commit()
    conn.close()

def get_files_list(db_conn):
    cursor = db_conn.cursor()
    cursor.execute("SELECT id, file_name FROM xlsx_files ORDER BY id")
    return cursor.fetchall()


def decode_koi8r_if_needed(value):
    if isinstance(value, bytes):
        try:
            return value.decode('koi8-r')
        except UnicodeDecodeError:
            return value.decode('utf-8', errors='replace')
    elif isinstance(value, str):
        try:
            return value.encode('cp1252').decode('koi8-r')
        except (UnicodeEncodeError, UnicodeDecodeError):
            return value
    else:
        return value

def decode_key(key):
    """
    Декодирует ключ (название столбца) из байтов или некорректной строки.
    Если ключ - str, пытаемся привести к нормальному виду, иначе возвращаем как есть.
    """
    if isinstance(key, bytes):
        try:
            return key.decode('koi8-r')
        except UnicodeDecodeError:
            return key.decode('utf-8', errors='replace')
    elif isinstance(key, str):
        try:
            return key.encode('cp1252').decode('koi8-r')
        except (UnicodeEncodeError, UnicodeDecodeError):
            return key
    else:
        return key


class ExportSelectedFilesThread(QThread):
    finished = pyqtSignal(str)

    def __init__(self, file_ids):
        super().__init__()
        self.file_ids = file_ids

    def run(self):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        for file_id in self.file_ids:
            cursor.execute("SELECT file_name FROM xlsx_files WHERE id = ?", (file_id,))
            file_name_row = cursor.fetchone()
            if not file_name_row:
                continue

            file_name = file_name_row[0]

            cursor.execute("""
                SELECT serial_number, time, task_number, message_type, data_length, data_blob, developer_note
                FROM messages WHERE xlsx_file_id = ?
                ORDER BY serial_number ASC
            """, (file_id,))
            messages = cursor.fetchall()

            if messages:
                wb = Workbook()
                ws = wb.active
                ws.title = "Messages"

                header = [
                    "Название файла", "Порядковый номер", "Время", "Номер задачи",
                    "Тип диагностического сообщения", "Длина бинарных данных",
                    "Бинарные данные", "Сообщение разработчику"
                ]
                ws.append(header)

                for msg in messages:
                    row = [file_name] + list(msg)
                    ws.append(row)


                ws.column_dimensions['A'].width = 20
                for col_idx, width in manual_widths.items():
                    excel_col_idx = col_idx + 1
                    col_letter = get_column_letter(excel_col_idx)
                    ws.column_dimensions[col_letter].width = width


                if not os.path.exists(EXPORT_DIR):
                    os.makedirs(EXPORT_DIR)

                # Защита имени файла от запрещённых символов
                safe_file_name = "".join(c for c in file_name if c.isalnum() or c in (' ', '_', '-')).rstrip()
                export_path = os.path.join(EXPORT_DIR, f"{safe_file_name}_exported.xlsx")
                wb.save(export_path)

        conn.close()
        self.finished.emit("Выбранные файлы экспортированы.")



class ExportEntireDatabaseThread(QThread):
    finished = pyqtSignal(str)

    def run(self):
        if not os.path.exists(EXPORT_DIR):
            os.makedirs(EXPORT_DIR)
        export_path = os.path.join(EXPORT_DIR, "full_database_export.sqlite")
        shutil.copyfile("database/converter.db", export_path)
        self.finished.emit("Вся база данных экспортирована.")


def load_excel_messages(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    messages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        msg = dict(zip(headers, row))
        messages.append(msg)
    return messages

def load_db_messages(file_id, db_connection):
    cursor = db_connection.cursor()
    cursor.execute("SELECT * FROM messages WHERE xlsx_file_id = ?", (file_id,))
    
    # Декодируем названия столбцов
    raw_columns = [desc[0] for desc in cursor.description]
    columns = [decode_key(c) for c in raw_columns]
    
    results = []
    for row in cursor.fetchall():
        msg = {}
        for i, val in enumerate(row):
            key = columns[i]
            msg[key] = decode_koi8r_if_needed(val)
        results.append(msg)
    return results

def compare_messages(excel_msgs, db_msgs):
    diffs = []
    stats = {
        "total_rows": 0,
        "diff_rows": 0,
        "matched_rows": 0,
        "field_diff_counts": {}
    }

    for i, (excel_msg, db_msg) in enumerate(zip(excel_msgs, db_msgs)):
        diff_fields = []
        for excel_field in FIELD_MAPPING:
            db_field = FIELD_MAPPING[excel_field]
            excel_value = str(excel_msg.get(excel_field, "")).strip()
            db_value = str(db_msg.get(db_field, "")).strip()
            if excel_value != db_value:
                diff_fields.append(excel_field)
                stats["field_diff_counts"][excel_field] = stats["field_diff_counts"].get(excel_field, 0) + 1

        if diff_fields:
            diffs.append({
                "row": i + 1,
                "diff_fields": diff_fields,
                "db_values": {FIELD_MAPPING[k]: db_msg.get(FIELD_MAPPING[k], "<нет>") for k in diff_fields},
                "excel_values": {k: excel_msg.get(k, "<нет>") for k in diff_fields}
            })

    stats["total_rows"] = len(excel_msgs)
    stats["diff_rows"] = len(diffs)
    stats["matched_rows"] = stats["total_rows"] - stats["diff_rows"]
    return diffs, stats


class CompareWorker(QThread):
    finished = pyqtSignal(list, dict)  # diffs, stats
    error = pyqtSignal(str)

    def __init__(self, excel_path, db_file_id):
        super().__init__()
        self.excel_path = excel_path
        self.db_file_id = db_file_id

    def run(self):
        try:
            db_conn = sqlite3.connect(DB_PATH)  # создаём здесь соединение

            excel_msgs = load_excel_messages(self.excel_path)
            db_msgs = load_db_messages(self.db_file_id, db_conn)
            diffs, stats = compare_messages(excel_msgs, db_msgs)

            db_conn.close()

            self.finished.emit(diffs, stats)
        except Exception as e:
            self.error.emit(str(e))


REVERSE_FIELD_MAPPING = {v: k for k, v in FIELD_MAPPING.items()}

class FilterSearchWorker(QThread):
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, db_path, file_id, filters, search_text):
        super().__init__()
        self.db_path = db_path
        self.file_id = file_id
        self.filters = filters
        self.search_text = search_text.lower() if search_text else None

    def run(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
    
            sql = """
                SELECT serial_number, time, task_number, message_type, data_length, data_blob, developer_note
                FROM messages
                WHERE xlsx_file_id = ?
            """
            params = [self.file_id]
    
            conditions = []
            if 'message_type' in self.filters and self.filters['message_type'] is not None:
                conditions.append("message_type = ?")
                params.append(self.filters['message_type'])
    
            if 'task_number_range' in self.filters:
                start, end = self.filters['task_number_range']
                conditions.append("task_number BETWEEN ? AND ?")
                params.extend([start, end])
    
            if conditions:
                sql += " AND " + " AND ".join(conditions)
    
            sql += " ORDER BY serial_number ASC"
    
            cursor.execute(sql, params)
            rows = cursor.fetchall()
    
            # Поля в порядке FIELD_MAPPING
            results = []
            for i, row in enumerate(rows, 1):
                serial_number, time_val, task_number, message_type, data_length, data_blob, developer_note = row
    
                msg = {
                    "Порядковый номер": i,
                    "Время": decode_koi8r_if_needed(time_val),
                    "Номер задачи": task_number,
                    "Тип диагностического сообщения": type_names.get(message_type, str(message_type)),
                    "Длина бинарных данных": data_length,
                    "Бинарные данные": decode_koi8r_if_needed(data_blob),
                    "Текстовое сообщение разработчику": decode_koi8r_if_needed(developer_note),
                }
    
                results.append(msg)
    
            # Фильтрация по тексту
            if self.search_text:
                filtered = []
                for msg in results:
                    if any(self.search_text in str(v).lower() for v in msg.values() if isinstance(v, str)):
                        filtered.append(msg)
                results = filtered
    
            self.finished.emit(results)
    
        except Exception as e:
            self.error.emit(str(e))
    
    

class MetabaseLauncherThread(QThread):
    started_successfully = pyqtSignal(bool)

    def __init__(self, jar_path, java_path, url):
        super().__init__()
        self.jar_path = jar_path
        self.java_path = java_path
        self.url = url
        self.process = None

    def is_metabase_running(self):
        try:
            requests.get(self.url, timeout=2)
            return True
        except:
            return False

    def run(self):
        if not os.path.exists(self.jar_path):
            self.started_successfully.emit(False)
            return

        try:
            self.process = subprocess.Popen(
                f'"{self.java_path}" -jar "{self.jar_path}"',
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                shell=True
            )
        except Exception:
            self.started_successfully.emit(False)
            return

        for _ in range(10):
            if self.is_metabase_running():
                self.started_successfully.emit(True)
                return
            time.sleep(1)

        self.started_successfully.emit(False)