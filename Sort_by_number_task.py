import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Global import sort_by_number_task_widths

def get_headers_from_wb(wb):
    for sheet in wb.worksheets:
        if sheet.max_row >= 1:
            headers = [cell.value for cell in sheet[1]]
            if headers:
                return headers
    return None

def gain_task_number(xlsx_file):
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    tasks = set()
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3 and row[2] is not None:
                tasks.add(str(row[2]).strip())
    return sorted(tasks, key=lambda x: int(x) if x.isdigit() else x)

def filter_rows_by_task(xlsx_file, task_col_index, selected_tasks):
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    header = None
    for sheet in wb.worksheets:
        if sheet.max_row >= 1:
            header = [cell.value for cell in sheet[1]]
            if header:
                break
    if header is None:
        raise ValueError("Заголовок не найден в файле")

    new_header = header[:task_col_index-1] + header[task_col_index:]
    data_by_task = {task: [] for task in selected_tasks}

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= task_col_index and row[task_col_index-1] is not None:
                task_val = str(row[task_col_index-1]).strip()
                if task_val in selected_tasks:
                    new_row = list(row[:task_col_index-1]) + list(row[task_col_index:])
                    data_by_task[task_val].append(new_row)

    return new_header, data_by_task

def create_sorted_workbook(new_header, data_by_task):
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for task, rows in data_by_task.items():
        sheet_title = f"{task} задача"
        ws = wb_out.create_sheet(title=sheet_title)
        ws.append(new_header)

        for r in rows:
            ws.append(r)
            num_cols = len(new_header)

        for i in range(1, num_cols + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = sort_by_number_task_widths.get(i, 10)

    return wb_out
