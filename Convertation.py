import struct
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from Global import manual_widths
import os
from database.db_manager import get_unique_file_name, insert_file, insert_messages

def parse_bin_file(bin_file):
    base_name = os.path.basename(bin_file)
    unique_name = get_unique_file_name(base_name)
    file_id = insert_file(unique_name)

    headers = ["Порядковый номер", "Время", "Номер задачи", "Тип диагностического сообщения",
               "Длина бинарных данных", "Бинарные данные", "Текстовое сообщение разработчику"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные 1"
    ws.append(headers)

    for i, width in manual_widths.items():
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    for col in range(4, 6):
        ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True)

    sheet_number = 1
    row_count = 1
    messages = []

    try:
        with open(bin_file, "rb") as f:
            while True:
                header = f.read(12)
                if not header or len(header) < 12:
                    break

                sequence_number, time_val, task_number, diag_type, bin_data_length = struct.unpack("<IIBBH", header)
                binary_data = f.read(bin_data_length) if bin_data_length > 0 else b""
                bin_data_hex = binary_data.hex().upper() if binary_data else "Отсутствует"

                dev_mess = bytearray()
                while True:
                    byte = f.read(1)
                    if not byte or byte == b'\0':
                        break
                    dev_mess.extend(byte)

                try:
                    dev_message = dev_mess.decode("koi8-r")
                except UnicodeDecodeError:
                    dev_message = dev_mess.decode("koi8-r", errors="replace")

                ws.append([sequence_number, time_val, task_number, diag_type, bin_data_length, bin_data_hex, dev_message])

                messages.append({
                    'serial_number': sequence_number,
                    'time': time_val,
                    'message_type': diag_type,
                    'task_number': task_number,
                    'data_length': bin_data_length,
                    'data_blob': bin_data_hex,
                    'developer_note': dev_message
                })

                row_count += 1
                if row_count > 10000:
                    sheet_number += 1
                    ws = wb.create_sheet(title=f"Данные {sheet_number}")
                    ws.append(headers)
                    for i, width in manual_widths.items():
                        col_letter = get_column_letter(i)
                        ws.column_dimensions[col_letter].width = width
                    for col in range(4, 6):
                        ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True)
                    row_count = 1

        # После обработки всех сообщений — вставка в базу
        insert_messages(file_id, messages)
        print(f"Добавляем сообщения: {len(messages)} штук")


    except Exception as e:
        print(f"Ошибка при чтении BIN файла: {e}")
        return None

    # Выравнивание в финале
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    return wb
