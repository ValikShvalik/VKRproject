import openpyxl
from openpyxl.utils import get_column_letter
from Global import sort_by_diag_type_widths, type_names

def sort_by_diag_type_message(xlsx_file, selected_types):
    wb_in = openpyxl.load_workbook(xlsx_file)
    sheet_in = wb_in.active

    headers = ["Порядковый номер", "Время", "Номер задачи",
               "Длина бинарных данных", "Бинарные данные", "Текстовое сообщение разработчику"]

    type_col_index = None
    for col_num, cell in enumerate(sheet_in[1], start=1):
        if cell.value == "Тип диагностического сообщения":
            type_col_index = col_num
            break
    if type_col_index is None:
        print("Ошибка: нет столбца 'Тип диагностического сообщения'")
        return None  # Ошибка: столбец не найден

    # Создаём словарь для хранения данных по типам сообщений
    data_by_type = {}
    
    for sheet_name in wb_in.sheetnames:
        sheet_in = wb_in[sheet_name]

        for row in sheet_in.iter_rows(min_row=2, values_only=True):
            msg_type = row[type_col_index - 1]
            if msg_type == 255:
                continue  # Пропускаем тип 255
            if msg_type not in data_by_type:
                data_by_type[msg_type] = []
            filtered_row = list(row)
            del filtered_row[type_col_index - 1]
            data_by_type[msg_type].append(filtered_row)

    # Создаём новый Workbook для сортированного файла
    wb_out = openpyxl.Workbook()
    first_sheet = True

    for msg_type in sorted(selected_types, reverse=True):
        if msg_type not in data_by_type:
            continue

        sheet_number = 1
        record_count = 0
        ws = None
        for row in data_by_type[msg_type]:
            if record_count % 500000 == 0:
                sheet_title = f"{type_names.get(msg_type, 'Неизвестный')}"
                if not first_sheet:
                    ws = wb_out.create_sheet(title=sheet_title)
                else:
                    ws = wb_out.active
                    ws.title = sheet_title
                    first_sheet = False
                ws.append(headers)
                for col_index, width in sort_by_diag_type_widths.items():
                    col_letter = get_column_letter(col_index)
                    ws.column_dimensions[col_letter].width = width
                sheet_number += 1
            ws.append(row)
            record_count += 1

    return wb_out 
